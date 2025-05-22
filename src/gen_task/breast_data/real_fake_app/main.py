import os
import random
from PIL import Image, ImageTk
import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class ImageGuessingApp:
    def __init__(self, root, image_dirs):
        self.root = root
        self.root.title("Visual Turing Test")
        self.root.geometry("800x600")  # Set window size

        self.image_dirs = image_dirs
        self.image_files = []
        self.image_labels = []
        self.current_image_index = 0
        self.current_set = None
        self.guess_results = []

        self.create_home_page()

    def create_home_page(self):
        self.clear_window()

        self.label = tk.Label(self.root, text="Selecionar Set", font=("Arial", 24))
        self.label.pack(pady=20)

        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=20)

        for set_name, dir_path in self.image_dirs.items():
            button = tk.Button(button_frame, text=set_name, command=lambda d=dir_path, s=set_name: self.start_set(d, s), font=("Arial", 16), width=20, height=2)
            button.pack(pady=10)

    def start_set(self, dir_path, set_name):
        self.image_files = []
        self.image_labels = []
        self.current_image_index = 0
        self.current_set = set_name
        self.guess_results = []

        self.load_images(dir_path)
        self.clear_window()
        self.create_game_page()

    def create_game_page(self):
        self.label = tk.Label(self.root)
        self.label.pack(pady=20)

        button_frame = tk.Frame(self.root)
        button_frame.pack(pady=20)

        self.real_button = tk.Button(button_frame, text="Real", command=lambda: self.check_guess('real'), font=("Arial", 16), width=10, height=2, bg="lightgreen")
        self.real_button.pack(side=tk.LEFT, padx=10)

        self.fake_button = tk.Button(button_frame, text="Falsa", command=lambda: self.check_guess('generated'), font=("Arial", 16), width=10, height=2, bg="lightcoral")
        self.fake_button.pack(side=tk.RIGHT, padx=10)

        self.status_label = tk.Label(self.root, text="", font=("Arial", 16))
        self.status_label.pack(pady=10)

        return_button_frame = tk.Frame(self.root)
        return_button_frame.pack(pady=20)

        return_button = tk.Button(return_button_frame, text="Regressar", command=self.create_home_page, font=("Arial", 14), width=10, height=2, bg="lightblue")
        return_button.pack(pady=10)

        self.show_next_image()

    def clear_window(self):
        for widget in self.root.winfo_children():
            widget.destroy()

    def load_images(self, dir_path):
        for root, _, files in os.walk(dir_path):
            for file in files:
                if file.endswith(('.png', '.jpg', '.jpeg')):
                    self.image_files.append(os.path.join(root, file))
                    label = 'real' if 'real' in file.lower() else 'generated'
                    self.image_labels.append(label)

    def show_next_image(self):
        if self.current_image_index < len(self.image_files):
            image_path = self.image_files[self.current_image_index]
            print(f"Showing image: {image_path}")  # Print the image path to the terminal

            image = Image.open(image_path)
            image = image.resize((400, 400), Image.ANTIALIAS)
            photo = ImageTk.PhotoImage(image)

            self.label.config(image=photo)
            self.label.image = photo

            self.status_label.config(text=f"Image {self.current_image_index + 1} of {len(self.image_files)}")
        else:
            self.clear_window()
            self.create_return_page()

    def check_guess(self, guess):
        correct_label = self.image_labels[self.current_image_index]
        result = "Correct" if guess == correct_label else f"Incorrect. It was {correct_label}."
        self.guess_results.append((self.image_files[self.current_image_index], guess, result))
        
        self.current_image_index += 1
        self.show_next_image()

    def create_return_page(self):
        self.label = tk.Label(self.root, text="Game Over", font=("Arial", 24))
        self.label.pack(pady=20)

        self.status_label = tk.Label(self.root, text=f"Results saved to {self.current_set}_results.xlsx", font=("Arial", 16))
        self.status_label.pack(pady=10)

        self.save_results_to_excel()

        return_button_frame = tk.Frame(self.root)
        return_button_frame.pack(pady=20)

        return_button = tk.Button(return_button_frame, text="Return", command=self.create_home_page, font=("Arial", 14), width=10, height=2, bg="lightblue")
        return_button.pack(pady=10)

    def save_results_to_excel(self):
        excel_filename = f"{self.current_set}_results.xlsx"
        if os.path.exists(excel_filename):
            wb = load_workbook(excel_filename)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Image Path", "Guess", "Result"])

        for image_path, guess, result in self.guess_results:
            ws.append([image_path, guess, result])

        wb.save(excel_filename)

if __name__ == "__main__":
    input_directories = {
        "Set 1": r"C:\Users\Diogo Campas\Desktop\Dissertação_23_24\code\VTT\Set_1",
        "Set 2": r"C:\Users\Diogo Campas\Desktop\Dissertação_23_24\code\VTT\Set_2",
        "Set 3": r"C:\Users\Diogo Campas\Desktop\Dissertação_23_24\code\VTT\Set_3"
    }

    root = tk.Tk()
    app = ImageGuessingApp(root, input_directories)
    root.mainloop()
